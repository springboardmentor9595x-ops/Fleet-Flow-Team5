"""
Reports & Export Router
Handles Fleet Utilization, Fuel Consumption, Driver Performance,
Delivery Performance, and Maintenance reports with PDF and Excel export capabilities.
"""
import io
from datetime import date, datetime, timedelta
from typing import Optional, List, Dict, Any
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, or_

from app.database import get_db
from app.core.deps import get_current_user
from app.models.user import User, RoleEnum
from app.models.vehicle import Vehicle
from app.models.driver import Driver
from app.models.trip import Trip
from app.models.shipment import Shipment, ShipmentStatusEnum
from app.models.fuel_record import FuelRecord
from app.models.maintenance import VehicleMaintenance
from app.models.attendance import Attendance

router = APIRouter()


# ── RBAC Scope Helper ────────────────────────────────────────────────────────
def check_report_permission(user: User, report_type: str):
    """
    Validates report access permissions based on user role:
    - Admin: All reports
    - FleetManager: All reports
    - Dispatcher: Delivery Performance only
    - Driver: Driver Performance only (personal summary)
    """
    if user.role == RoleEnum.Admin or user.role == RoleEnum.FleetManager:
        return True
    
    if user.role == RoleEnum.Dispatcher:
        if report_type != "delivery-performance":
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Dispatchers have operational access to Delivery Performance Reports only.",
            )
        return True

    if user.role == RoleEnum.Driver:
        if report_type != "driver-performance":
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Drivers can only view their own Driver Performance summary.",
            )
        return True

    raise HTTPException(
        status_code=status.HTTP_403_FORBIDDEN,
        detail="Unauthorized to access reports.",
    )


# ── 1. Fleet Utilization Report ──────────────────────────────────────────────
@router.get("/fleet-utilization")
def get_fleet_utilization_report(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    check_report_permission(current_user, "fleet-utilization")
    
    if not end_date:
        end_date = date.today()
    if not start_date:
        start_date = end_date - timedelta(days=30)

    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt = datetime.combine(end_date, datetime.max.time())

    vehicles = db.query(Vehicle).all()
    total_vehicles = len(vehicles)

    status_counts = {"Available": 0, "On Trip": 0, "Maintenance": 0, "Inactive": 0}
    for v in vehicles:
        st = v.status or "Available"
        status_counts[st] = status_counts.get(st, 0) + 1

    # Query trips in date range
    trips = db.query(Trip).filter(Trip.created_at >= start_dt, Trip.created_at <= end_dt).all()
    
    # Calculate vehicle breakdown
    vehicle_rows = []
    active_vehicles_count = 0
    for v in vehicles:
        v_trips = [t for t in trips if t.vehicle_id == v.vehicle_id]
        trip_count = len(v_trips)
        total_distance = sum(float(t.distance or 0) for t in v_trips)
        if trip_count > 0:
            active_vehicles_count += 1

        vehicle_rows.append({
            "vehicle_id": str(v.vehicle_id),
            "registration_number": v.registration_number,
            "vehicle_type": v.vehicle_type,
            "brand": v.brand or "—",
            "model": v.model or "—",
            "status": v.status or "Available",
            "trips_count": trip_count,
            "total_distance_km": round(total_distance, 1),
            "capacity": v.capacity or 0,
        })

    utilization_rate = round((active_vehicles_count / max(total_vehicles, 1)) * 100, 1)

    return {
        "report_title": "Fleet Utilization Report",
        "start_date": str(start_date),
        "end_date": str(end_date),
        "total_vehicles": total_vehicles,
        "active_vehicles": active_vehicles_count,
        "utilization_rate": utilization_rate,
        "total_trips_in_period": len(trips),
        "status_breakdown": status_counts,
        "data": vehicle_rows,
    }


# ── 2. Fuel Consumption Report ───────────────────────────────────────────────
@router.get("/fuel-consumption")
def get_fuel_consumption_report(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    check_report_permission(current_user, "fuel-consumption")

    if not end_date:
        end_date = date.today()
    if not start_date:
        start_date = end_date - timedelta(days=30)

    records = db.query(FuelRecord).filter(
        FuelRecord.refill_date >= start_date,
        FuelRecord.refill_date <= end_date,
    ).all()

    total_liters = sum(float(r.fuel_amount or 0) for r in records)
    total_cost = sum(float(r.fuel_cost or 0) for r in records)
    avg_cost_per_liter = round(total_cost / max(total_liters, 1), 2)

    # Group by vehicle
    vehicles_dict = {}
    for r in records:
        vid = str(r.vehicle_id)
        if vid not in vehicles_dict:
            veh = db.query(Vehicle).filter(Vehicle.vehicle_id == r.vehicle_id).first()
            vehicles_dict[vid] = {
                "vehicle_id": vid,
                "registration_number": veh.registration_number if veh else "Unknown",
                "vehicle_type": veh.vehicle_type if veh else "—",
                "refill_count": 0,
                "total_liters": 0.0,
                "total_cost": 0.0,
                "latest_odometer": 0.0,
            }
        vehicles_dict[vid]["refill_count"] += 1
        vehicles_dict[vid]["total_liters"] += float(r.fuel_amount or 0)
        vehicles_dict[vid]["total_cost"] += float(r.fuel_cost or 0)
        if float(r.mileage or 0) > vehicles_dict[vid]["latest_odometer"]:
            vehicles_dict[vid]["latest_odometer"] = float(r.mileage or 0)

    rows = []
    for vid, item in vehicles_dict.items():
        item["total_liters"] = round(item["total_liters"], 1)
        item["total_cost"] = round(item["total_cost"], 2)
        item["avg_cost_per_liter"] = round(item["total_cost"] / max(item["total_liters"], 1), 2)
        rows.append(item)

    rows.sort(key=lambda x: x["total_cost"], reverse=True)

    return {
        "report_title": "Fuel Consumption Report",
        "start_date": str(start_date),
        "end_date": str(end_date),
        "total_fuel_liters": round(total_liters, 1),
        "total_fuel_cost": round(total_cost, 2),
        "avg_cost_per_liter": avg_cost_per_liter,
        "total_refills": len(records),
        "data": rows,
    }


# ── 3. Driver Performance Report ─────────────────────────────────────────────
@router.get("/driver-performance")
def get_driver_performance_report(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    check_report_permission(current_user, "driver-performance")

    if not end_date:
        end_date = date.today()
    if not start_date:
        start_date = end_date - timedelta(days=30)

    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt = datetime.combine(end_date, datetime.max.time())

    # Scope driver query
    if current_user.role == RoleEnum.Driver:
        driver = db.query(Driver).filter(Driver.user_id == current_user.user_id).first()
        drivers = [driver] if driver else []
    else:
        drivers = db.query(Driver).all()

    trips = db.query(Trip).filter(Trip.created_at >= start_dt, Trip.created_at <= end_dt).all()
    attendance_records = db.query(Attendance).filter(
        Attendance.date >= start_date,
        Attendance.date <= end_date,
    ).all()

    rows = []
    total_fleet_trips = 0
    total_fleet_completed = 0

    for d in drivers:
        if not d:
            continue
        user = db.query(User).filter(User.user_id == d.user_id).first()
        name = user.full_name if user else "Driver"
        email = user.email if user else "—"

        d_trips = [t for t in trips if t.driver_id == d.driver_id]
        completed = [t for t in d_trips if t.status == "Completed"]
        delayed = [t for t in d_trips if t.status == "Delayed"]
        distance = sum(float(t.distance or 0) for t in d_trips)

        d_attendance = [a for a in attendance_records if a.driver_id == d.driver_id]
        present_days = len([a for a in d_attendance if a.status == "Present"])
        leave_days = len([a for a in d_attendance if a.status == "Leave"])
        absent_days = len([a for a in d_attendance if a.status == "Absent"])
        total_att = len(d_attendance)
        att_rate = round((present_days / max(total_att, 1)) * 100, 1) if total_att > 0 else 100.0

        trip_count = len(d_trips)
        completed_count = len(completed)
        total_fleet_trips += trip_count
        total_fleet_completed += completed_count

        on_time_rate = round((completed_count / max(trip_count, 1)) * 100, 1) if trip_count > 0 else 100.0

        rows.append({
            "driver_id": str(d.driver_id),
            "driver_name": name,
            "email": email,
            "license_number": d.license_number or "—",
            "status": d.status or "Active",
            "total_trips": trip_count,
            "completed_trips": completed_count,
            "delayed_trips": len(delayed),
            "on_time_rate_pct": on_time_rate,
            "total_distance_km": round(distance, 1),
            "attendance_present_days": present_days,
            "attendance_leave_days": leave_days,
            "attendance_absent_days": absent_days,
            "attendance_rate_pct": att_rate,
        })

    rows.sort(key=lambda x: x["total_trips"], reverse=True)

    return {
        "report_title": "Driver Performance Report",
        "start_date": str(start_date),
        "end_date": str(end_date),
        "total_drivers_analyzed": len(drivers),
        "total_trips": total_fleet_trips,
        "overall_completion_rate": round((total_fleet_completed / max(total_fleet_trips, 1)) * 100, 1),
        "data": rows,
    }


# ── 4. Delivery Performance Report ───────────────────────────────────────────
@router.get("/delivery-performance")
def get_delivery_performance_report(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    check_report_permission(current_user, "delivery-performance")

    if not end_date:
        end_date = date.today()
    if not start_date:
        start_date = end_date - timedelta(days=30)

    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt = datetime.combine(end_date, datetime.max.time())

    shipments = db.query(Shipment).filter(
        Shipment.created_at >= start_dt,
        Shipment.created_at <= end_dt,
    ).order_by(Shipment.created_at.desc()).all()

    total_shipments = len(shipments)
    delivered = [s for s in shipments if s.status == ShipmentStatusEnum.Delivered]
    delayed = [s for s in shipments if s.status == ShipmentStatusEnum.Delayed]
    in_transit = [s for s in shipments if s.status == ShipmentStatusEnum.InTransit]
    created = [s for s in shipments if s.status in (ShipmentStatusEnum.Created, ShipmentStatusEnum.Assigned)]
    cancelled = [s for s in shipments if s.status == ShipmentStatusEnum.Cancelled]

    total_weight = sum(float(s.shipment_weight or 0) for s in shipments)
    on_time_count = len(delivered) + len(in_transit) + len(created)
    on_time_rate = round(((total_shipments - len(delayed) - len(cancelled)) / max(total_shipments, 1)) * 100, 1)

    rows = []
    for s in shipments:
        st_val = s.status.value if hasattr(s.status, 'value') else str(s.status)
        rows.append({
            "shipment_id": str(s.shipment_id),
            "tracking_number": s.tracking_number,
            "customer_name": s.customer_name,
            "source": s.source,
            "destination": s.destination,
            "status": st_val,
            "weight_kg": float(s.shipment_weight or 0),
            "expected_delivery": s.expected_delivery.strftime("%Y-%m-%d %H:%M") if s.expected_delivery else "—",
            "created_at": s.created_at.strftime("%Y-%m-%d %H:%M"),
        })

    return {
        "report_title": "Delivery Performance Report",
        "start_date": str(start_date),
        "end_date": str(end_date),
        "total_shipments": total_shipments,
        "delivered_count": len(delivered),
        "delayed_count": len(delayed),
        "in_transit_count": len(in_transit),
        "cancelled_count": len(cancelled),
        "on_time_delivery_rate_pct": on_time_rate,
        "total_cargo_weight_kg": round(total_weight, 1),
        "data": rows,
    }


# ── 5. Maintenance Report ───────────────────────────────────────────────────
@router.get("/maintenance")
def get_maintenance_report(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    check_report_permission(current_user, "maintenance")

    if not end_date:
        end_date = date.today()
    if not start_date:
        start_date = end_date - timedelta(days=30)

    records = db.query(VehicleMaintenance).filter(
        VehicleMaintenance.service_date >= start_date,
        VehicleMaintenance.service_date <= end_date,
    ).order_by(VehicleMaintenance.service_date.desc()).all()

    total_cost = sum(float(r.cost or 0) for r in records)
    resolved_count = len([r for r in records if (r.status or "").lower() == "resolved"])
    scheduled_count = len([r for r in records if (r.status or "").lower() in ("scheduled", "in progress", "pending")])

    # Breakdown by maintenance type
    type_breakdown = {}
    for r in records:
        mtype = r.maintenance_type or "General Service"
        type_breakdown[mtype] = type_breakdown.get(mtype, 0) + 1

    # Group by vehicle
    veh_dict = {}
    for r in records:
        vid = str(r.vehicle_id) if r.vehicle_id else "unknown"
        if vid not in veh_dict:
            veh = db.query(Vehicle).filter(Vehicle.vehicle_id == r.vehicle_id).first() if r.vehicle_id else None
            veh_dict[vid] = {
                "vehicle_id": vid,
                "registration_number": veh.registration_number if veh else "Unknown",
                "service_count": 0,
                "total_cost": 0.0,
            }
        veh_dict[vid]["service_count"] += 1
        veh_dict[vid]["total_cost"] += float(r.cost or 0)

    rows = []
    for r in records:
        veh = db.query(Vehicle).filter(Vehicle.vehicle_id == r.vehicle_id).first() if r.vehicle_id else None
        rows.append({
            "maintenance_id": str(r.maintenance_id),
            "registration_number": veh.registration_number if veh else "Vehicle",
            "maintenance_type": r.maintenance_type,
            "service_date": str(r.service_date),
            "next_service_date": str(r.next_service_date) if r.next_service_date else "—",
            "cost": float(r.cost or 0),
            "status": r.status or "Scheduled",
            "remarks": r.remarks or "—",
        })

    return {
        "report_title": "Fleet Maintenance & Servicing Report",
        "start_date": str(start_date),
        "end_date": str(end_date),
        "total_records": len(records),
        "total_maintenance_expense": round(total_cost, 2),
        "resolved_services": resolved_count,
        "pending_scheduled_services": scheduled_count,
        "type_frequency": type_breakdown,
        "cost_by_vehicle": list(veh_dict.values()),
        "data": rows,
    }


# ── PDF Export Functionality ────────────────────────────────────────────────
@router.get("/export/pdf")
def export_report_pdf(
    report_type: str = Query(..., pattern="^(fleet-utilization|fuel-consumption|driver-performance|delivery-performance|maintenance)$"),
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    check_report_permission(current_user, report_type)

    if not end_date:
        end_date = date.today()
    if not start_date:
        start_date = end_date - timedelta(days=30)

    # Fetch corresponding report data
    if report_type == "fleet-utilization":
        data = get_fleet_utilization_report(start_date, end_date, db, current_user)
        headers = ["Reg Number", "Type", "Status", "Trips", "Distance (km)", "Capacity"]
        keys = ["registration_number", "vehicle_type", "status", "trips_count", "total_distance_km", "capacity"]
    elif report_type == "fuel-consumption":
        data = get_fuel_consumption_report(start_date, end_date, db, current_user)
        headers = ["Reg Number", "Type", "Refills", "Total Liters", "Total Cost ($)", "Avg $/L"]
        keys = ["registration_number", "vehicle_type", "refill_count", "total_liters", "total_cost", "avg_cost_per_liter"]
    elif report_type == "driver-performance":
        data = get_driver_performance_report(start_date, end_date, db, current_user)
        headers = ["Driver Name", "Status", "Trips", "Completed", "On-Time %", "Distance (km)", "Present Days", "Attendance %"]
        keys = ["driver_name", "status", "total_trips", "completed_trips", "on_time_rate_pct", "total_distance_km", "attendance_present_days", "attendance_rate_pct"]
    elif report_type == "delivery-performance":
        data = get_delivery_performance_report(start_date, end_date, db, current_user)
        headers = ["Tracking #", "Customer", "Source", "Destination", "Status", "Weight (kg)"]
        keys = ["tracking_number", "customer_name", "source", "destination", "status", "weight_kg"]
    else: # maintenance
        data = get_maintenance_report(start_date, end_date, db, current_user)
        headers = ["Reg Number", "Service Type", "Date", "Next Date", "Cost ($)", "Status"]
        keys = ["registration_number", "maintenance_type", "service_date", "next_service_date", "cost", "status"]

    # Generate PDF using ReportLab
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30,
    )
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Heading1"],
        fontSize=18,
        leading=22,
        textColor=colors.HexColor("#0F172A"),
        fontName="Helvetica-Bold",
    )
    sub_style = ParagraphStyle(
        "ReportSub",
        parent=styles["Normal"],
        fontSize=10,
        leading=14,
        textColor=colors.HexColor("#64748B"),
    )
    cell_style = ParagraphStyle(
        "Cell",
        parent=styles["Normal"],
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor("#1E293B"),
    )
    header_cell_style = ParagraphStyle(
        "HeaderCell",
        parent=styles["Normal"],
        fontSize=9,
        leading=12,
        textColor=colors.white,
        fontName="Helvetica-Bold",
    )

    elements = []
    elements.append(Paragraph(f"FleetFlow Logistics — {data['report_title']}", title_style))
    elements.append(Spacer(1, 4))
    elements.append(Paragraph(f"Period: {start_date} to {end_date} | Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')} | Requested by: {current_user.full_name} ({current_user.role.value if hasattr(current_user.role, 'value') else current_user.role})", sub_style))
    elements.append(Spacer(1, 14))

    # Summary box
    summary_text = ""
    for k, v in data.items():
        if k not in ("data", "status_breakdown", "type_frequency", "cost_by_vehicle", "report_title", "start_date", "end_date"):
            label = k.replace("_", " ").title()
            summary_text += f"<b>{label}:</b> {v} &nbsp;&nbsp;|&nbsp;&nbsp; "
    if summary_text:
        elements.append(Paragraph(f"<font color='#0D9488'><b>Summary Metrics:</b></font> {summary_text.rstrip('&nbsp;&nbsp;|&nbsp;&nbsp; ')}", sub_style))
        elements.append(Spacer(1, 14))

    # Build Table
    table_data = [[Paragraph(h, header_cell_style) for h in headers]]
    for item in data.get("data", [])[:200]: # max 200 rows for PDF layout
        row = [Paragraph(str(item.get(k, "—")), cell_style) for k in keys]
        table_data.append(row)

    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0D9488")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
        ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#CBD5E1")),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    filename = f"FleetFlow_{report_type.replace('-', '_')}_{start_date}_{end_date}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── Excel Export Functionality ──────────────────────────────────────────────
@router.get("/export/excel")
def export_report_excel(
    report_type: str = Query(..., pattern="^(fleet-utilization|fuel-consumption|driver-performance|delivery-performance|maintenance)$"),
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    check_report_permission(current_user, report_type)

    if not end_date:
        end_date = date.today()
    if not start_date:
        start_date = end_date - timedelta(days=30)

    if report_type == "fleet-utilization":
        data = get_fleet_utilization_report(start_date, end_date, db, current_user)
        headers = ["Registration #", "Vehicle Type", "Status", "Trips Count", "Total Distance (km)", "Capacity"]
        keys = ["registration_number", "vehicle_type", "status", "trips_count", "total_distance_km", "capacity"]
    elif report_type == "fuel-consumption":
        data = get_fuel_consumption_report(start_date, end_date, db, current_user)
        headers = ["Registration #", "Vehicle Type", "Refill Count", "Total Liters", "Total Cost ($)", "Avg $/L", "Latest Odometer (km)"]
        keys = ["registration_number", "vehicle_type", "refill_count", "total_liters", "total_cost", "avg_cost_per_liter", "latest_odometer"]
    elif report_type == "driver-performance":
        data = get_driver_performance_report(start_date, end_date, db, current_user)
        headers = ["Driver Name", "Email", "License #", "Status", "Total Trips", "Completed Trips", "Delayed Trips", "On-Time %", "Distance (km)", "Present Days", "Leave Days", "Absent Days", "Attendance Rate %"]
        keys = ["driver_name", "email", "license_number", "status", "total_trips", "completed_trips", "delayed_trips", "on_time_rate_pct", "total_distance_km", "attendance_present_days", "attendance_leave_days", "attendance_absent_days", "attendance_rate_pct"]
    elif report_type == "delivery-performance":
        data = get_delivery_performance_report(start_date, end_date, db, current_user)
        headers = ["Tracking #", "Customer", "Source", "Destination", "Status", "Weight (kg)", "Expected Delivery", "Created At"]
        keys = ["tracking_number", "customer_name", "source", "destination", "status", "weight_kg", "expected_delivery", "created_at"]
    else: # maintenance
        data = get_maintenance_report(start_date, end_date, db, current_user)
        headers = ["Registration #", "Service Type", "Service Date", "Next Recurring Date", "Cost ($)", "Status", "Remarks"]
        keys = ["registration_number", "maintenance_type", "service_date", "next_service_date", "cost", "status", "remarks"]

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = data["report_title"][:31]

    # Title & Metadata
    title_font = Font(name="Segoe UI", size=16, bold=True, color="0D9488")
    meta_font = Font(name="Segoe UI", size=10, italic=True, color="64748B")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    data_font = Font(name="Segoe UI", size=10)
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    ws["A1"] = f"FleetFlow Logistics — {data['report_title']}"
    ws["A1"].font = title_font
    ws["A2"] = f"Date Range: {start_date} to {end_date} | Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')} | Requested By: {current_user.full_name}"
    ws["A2"].font = meta_font

    # Summary Row
    summary_items = [f"{k.replace('_', ' ').title()}: {v}" for k, v in data.items() if k not in ("data", "status_breakdown", "type_frequency", "cost_by_vehicle", "report_title", "start_date", "end_date")]
    if summary_items:
        ws["A3"] = " | ".join(summary_items)
        ws["A3"].font = Font(name="Segoe UI", size=10, bold=True, color="0F766E")

    start_row = 5
    # Write Header
    for col_num, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Write Data
    for row_idx, item in enumerate(data.get("data", []), start_row + 1):
        for col_idx, k in enumerate(keys, 1):
            val = item.get(k, "—")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")

    # Auto-adjust column width
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row >= start_row and cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"FleetFlow_{report_type.replace('-', '_')}_{start_date}_{end_date}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
