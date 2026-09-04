import io
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def generate_excel_report(report_data: dict) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active

    title = report_data.get("title", "FleetFlow Report")
    ws.title = "FleetFlow Report"

    # Styling Definitions
    title_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")

    subtitle_font = Font(name="Calibri", size=10, italic=True, color="64748B")

    section_font = Font(name="Calibri", size=12, bold=True, color="0F172A")
    section_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

    label_font = Font(name="Calibri", size=10, bold=True, color="334155")
    val_font = Font(name="Calibri", size=10, color="0F172A")

    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")

    row_fill_even = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    row_fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    # 1. Title Banner (Merged A1:F1)
    columns = report_data.get("columns", [])
    max_col_idx = max(len(columns), 4)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col_idx)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 32

    # 2. Metadata Subtitle
    generated_at = report_data.get("generated_at", "")
    period = report_data.get("period", {})
    start_date = period.get("start_date", "All Time")
    end_date = period.get("end_date", "Present")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col_idx)
    sub_cell = ws.cell(row=2, column=1, value=f"Generated: {generated_at}  |  Period: {start_date} to {end_date}")
    sub_cell.font = subtitle_font
    ws.row_dimensions[2].height = 18

    current_row = 4

    # 3. Executive Summary Block
    summary_items = report_data.get("summary", [])
    if summary_items:
        ws.cell(row=current_row, column=1, value="Executive Summary").font = section_font
        current_row += 1

        for item in summary_items:
            c1 = ws.cell(row=current_row, column=1, value=item["label"])
            c1.font = label_font
            c1.fill = section_fill
            c1.border = thin_border

            c2 = ws.cell(row=current_row, column=2, value=str(item["value"]))
            c2.font = val_font
            c2.border = thin_border
            current_row += 1

        current_row += 1

    # 4. Detailed Data Table
    rows = report_data.get("rows", [])
    if columns:
        ws.cell(row=current_row, column=1, value="Detailed Records").font = section_font
        current_row += 1

        # Header Row
        ws.row_dimensions[current_row].height = 24
        for col_idx, col_name in enumerate(columns, start=1):
            h_cell = ws.cell(row=current_row, column=col_idx, value=col_name)
            h_cell.font = header_font
            h_cell.fill = header_fill
            h_cell.alignment = Alignment(horizontal="center", vertical="center")
            h_cell.border = thin_border

        current_row += 1

        # Data Rows
        for r_idx, row_data in enumerate(rows):
            ws.row_dimensions[current_row].height = 20
            row_fill = row_fill_even if r_idx % 2 == 1 else row_fill_odd
            for col_idx, val in enumerate(row_data, start=1):
                d_cell = ws.cell(row=current_row, column=col_idx, value=str(val))
                d_cell.font = val_font
                d_cell.fill = row_fill
                d_cell.border = thin_border
                d_cell.alignment = Alignment(horizontal="left", vertical="center")
            current_row += 1

    # 5. Auto-fit column widths
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for c in col:
            val_str = str(c.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    buffer = io.BytesIO()
    wb.save(buffer)
    excel_bytes = buffer.getvalue()
    buffer.close()
    return excel_bytes
